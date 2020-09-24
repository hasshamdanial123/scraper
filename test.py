from selenium import webdriver
import pandas as pd
import openpyxl
path=("C:/Users/hp/Desktop/web-automation/abc.xlsx")
Workbook=openpyxl.load_workbook(path)
sheet=Workbook.active
for r in range(4,6):
    for c in range(4,8):
        sheet.cell(row=r,column=c).value="asjbnn"
Workbook.save(path)
df = pd.read_excel('abc.xlsx')
sheet_years = df['candidate_name']
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
import time
from webdriver_manager.chrome import ChromeDriverManager
j,i = 2, 5 
for year in sheet_years:
    print(year)
    browser = webdriver.Chrome(ChromeDriverManager().install())
    browser.get('https://www.electoralcommission.org.uk/2019-candidate-spending')
    hs = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[3]/div[2]/div[1]/div[1]/div/form/input").send_keys(year)
    time.sleep(3)
    name = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[3]/div[2]/div[1]/div[1]/div/ul/li").click()
    time.sleep(2)
    browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[3]/div/button").click()
    #browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[3]/div/div/div/button[4]").click()
    abc = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[3]/div/div/div/button[5]")
    #browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[3]/div/div/div/button[1]").click()
    #browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[3]/div/div/div/button[2]").click()
    element = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[1]/div[2]/div/div/div/div")
    price = element.get_attribute('innerHTML')
    h1=price.strip()
    sheet.cell(row=j,column=i).value=h1
    Workbook.save(path)
    i=i+1
    element2 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[2]/div[2]/div/div/div/div")
    price2 = element2.get_attribute('innerHTML')
    h2=price2.strip()
    sheet.cell(row=j,column=i).value=h2
    Workbook.save(path)
    i=i+1
    element3 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[3]/div[2]/div/div/div/div")
    price3 = element3.get_attribute('innerHTML')
    h3=price3.strip()
    sheet.cell(row=j,column=i).value=h3
    Workbook.save(path)
    i=i+1
    element4 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[4]/div[2]/div/div/div/div")
    price4 = element4.get_attribute('innerHTML')
    h4=price4.strip()
    sheet.cell(row=j,column=i).value=h4
    i=i+1
    Workbook.save(path)
    element5 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[5]/div[2]/div/div/div/div")
    price5 = element5.get_attribute('innerHTML')
    h5=price5.strip()
    sheet.cell(row=j,column=i).value=h5
    Workbook.save(path)
    i=i+1
    element6 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[6]/div[2]/div/div/div/div")
    price6 = element6.get_attribute('innerHTML')
    h6=price6.strip()
    sheet.cell(row=j,column=i).value=h6
    Workbook.save(path)
    i=i+1
    element7 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[7]/div[2]/div/div/div/div")
    price7 = element7.get_attribute('innerHTML')
    h7=price7.strip()
    sheet.cell(row=j,column=i).value=h7
    Workbook.save(path)
    i=i+1
    element8 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[8]/div[2]/div/div/div/div")
    price8 = element8.get_attribute('innerHTML')
    h8=price8.strip()
    sheet.cell(row=j,column=i).value=h8
    Workbook.save(path)
    i=i+1
    element9 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[9]/div[2]/div/div/div/div")
    price9 = element9.get_attribute('innerHTML')
    h9=price9.strip()
    sheet.cell(row=j,column=i).value=h9
    Workbook.save(path)
    i=i+1
    element10 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[9]/div[2]/div/div/div/div")
    price10 = element10.get_attribute('innerHTML')
    h10=price10.strip()
    sheet.cell(row=j,column=i).value=h10
    Workbook.save(path)
    i=i+1
    element11 = browser.find_element_by_xpath("/html/body/div[2]/div[2]/div/main/article/div[3]/div[2]/div[2]/div/div/div/div[1]/section/div/div/div/div/div/div[4]/div/div[10]/div[2]/div/div/div/div")
    price11 = element11.get_attribute('innerHTML')
    h11=price11.strip()
    sheet.cell(row=j,column=i).value=h11
    Workbook.save(path)
    j=j+1
    i=5
    print(abc)
    print(name)